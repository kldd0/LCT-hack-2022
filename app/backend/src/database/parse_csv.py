import openpyxl
from pydantic import BaseModel, validator
from typing import Mapping, List
import json
from dataclasses import dataclass


class Info(BaseModel):
    type_name: str
    name: str
    address: Mapping[str, str]
    district: str


@dataclass
class Req:
    street: str
    house: str
    country: str = "Российская Федерация"
    city: str = "Москва"

    @property
    def parse(self, addr: Mapping[str, str]) -> None:
        for key, val in addr:
            if key == "house":
                self.city = val
            elif key == "street":
                self.street = val

    @property
    def res(self) -> str:
        return f'''https://maps.mail.ru/osm/tools/overpass/api/interpreter?data=[out:json];nwr["addr:street"="{self.street}"]["addr:housenumber"="{self.house}"];out;'''
        


wookbook = openpyxl.load_workbook("data-104108-2022-10-23.xlsx")

worksheet = wookbook.active


def parse_address(addr: List[str]) -> Mapping[str, str]:
    result: Mapping[str, str] = dict()
    for v in addr:
        if "Российская Федерация" in v:
            result["country"] = v
        elif "Москва" in v:
            result["city"] = v
        elif "округ" in v:
            continue
        elif "дом" in v:
            result["house"] = "".join(v.split()[1:])
        elif "строение" in v:
            result["house"] += " с" + "".join(v.split()[1:])
        else:
            result["street"] = v
    return result


def request_template():
    s = f'''https://maps.mail.ru/osm/tools/overpass/api/interpreter?data=[out:json];nwr["addr:street"="{addr['street']}"]["addr:housenumber"="{addr['house']}"];out;'''


indexes = [8, 2, 5, 4] # ? i: 3
col_names = ["type_name", "name", "address", "district"]
for rowi in range(3, 4):
    obj = dict()
    for coli, name in zip(indexes, col_names):
        if name == "address":
            obj[name] = parse_address(worksheet.cell(row=rowi, column=coli).value.split(', '))
        else:
            obj[name] = worksheet.cell(row=rowi, column=coli).value
    obj = Info.parse_obj(obj)
    with open("test.json", "w") as file:
        file.write(json.dumps(obj.dict(), indent=4, ensure_ascii=False).encode("utf-8").decode())
    print("")